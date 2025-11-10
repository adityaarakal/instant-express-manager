#!/usr/bin/env python3
"""
Export the `Planned Expenses` worksheet into structured JSON seed data.

Usage:
    python scripts/export_planned_expenses.py \
        --workbook docs/Copy\ of\ Expense\ Shares.xlsx \
        --output data/seeds/planned-expenses.json

The script mirrors the repeating monthly blocks in the sheet and produces an
array of months, each containing metadata (fixed factor, due dates, statuses)
and a list of per-account allocations. The goal is to provide high-fidelity
seed data for the React application while staying faithful to the Excel logic.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Set

ROOT_DIR = Path(__file__).resolve().parents[1]
PYDEPS_DIR = ROOT_DIR / ".pydeps"

if PYDEPS_DIR.exists():
    sys.path.insert(0, str(PYDEPS_DIR))

try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install it via `python3 -m pip install openpyxl "
        "--target ./.pydeps`."
    ) from exc


# --------------------------------------------------------------------------- #
# Data structures
# --------------------------------------------------------------------------- #


Number = Optional[float]


def _to_number(value) -> Number:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _to_iso_date(value) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


@dataclass
class AccountAllocation:
    name: str
    remaining_cash: Number
    fixed_balance: Number
    savings_transfer: Number
    bucket_allocations: Dict[str, Number] = field(default_factory=dict)
    formulas: Dict[str, Optional[str]] = field(default_factory=dict)
    bucket_formulas: Dict[str, Optional[str]] = field(default_factory=dict)


@dataclass
class PlannedMonthExport:
    month_start: str
    fixed_factor: Number
    inflow_total: Number
    inflow_formula: Optional[str]
    fixed_factor_formula: Optional[str]
    status_by_bucket: Dict[str, str]
    due_dates: Dict[str, Optional[str]]
    bucket_order: List[str]
    accounts: List[AccountAllocation]
    source_rows: Dict[str, int]
    ref_errors: List[Dict[str, Optional[str]]] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #


def _iter_month_starts(ws) -> Iterable[int]:
    """Yield sheet row indices that start a month block."""
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        header_flag = ws.cell(row=row, column=6).value  # often 'Balance'
        if isinstance(cell_value, datetime) and isinstance(header_flag, str):
            yield row


def _extract_bucket_columns(ws, legend_row: int) -> Dict[int, str]:
    """Return mapping of column index → bucket name for the block."""
    columns: Dict[int, str] = {}
    occurrences: Dict[str, int] = {}
    for col in range(4, 14):  # cover D–M
        header = ws.cell(row=legend_row, column=col).value
        if isinstance(header, str) and header.strip():
            base = header.strip()
            count = occurrences.get(base, 0)
            occurrences[base] = count + 1
            name = base if count == 0 else f"{base} ({get_column_letter(col)})"
            columns[col] = name
    return columns


def _extract_statuses(ws, status_row: int, bucket_cols: Dict[int, str]) -> Dict[str, str]:
    statuses: Dict[str, str] = {}
    for col, bucket in bucket_cols.items():
        raw = ws.cell(row=status_row, column=col).value
        statuses[bucket] = str(raw).strip() if isinstance(raw, str) and raw.strip() else "Pending"
    return statuses


def _inspect_cell(
    ws_values,
    ws_formulas,
    row: int,
    column: int,
    ref_errors: List[Dict[str, Optional[str]]],
    ref_seen: Set[str],
):
    value_cell = ws_values.cell(row=row, column=column)
    formula_cell = ws_formulas.cell(row=row, column=column)

    raw_value = value_cell.value
    formula: Optional[str] = None

    if formula_cell.data_type == "f" and formula_cell.value:
        formula_str = str(formula_cell.value)
        formula = formula_str if formula_str.startswith("=") else f"={formula_str}"
    elif formula_cell.data_type == "e" and formula_cell.value:
        formula = str(formula_cell.value)

    cell_address = f"{get_column_letter(column)}{row}"
    has_ref_error = False

    if isinstance(raw_value, str) and "#REF" in raw_value.upper():
        has_ref_error = True
    if formula and "#REF" in formula.upper():
        has_ref_error = True
    if formula_cell.data_type == "e":
        has_ref_error = True

    if has_ref_error and cell_address not in ref_seen:
        ref_seen.add(cell_address)
        ref_errors.append(
            {
                "cell": cell_address,
                "value": None if raw_value is None else str(raw_value),
                "formula": formula,
            }
        )

    return raw_value, formula


def _extract_due_dates(
    ws_values,
    ws_formulas,
    start_row: int,
    bucket_cols: Dict[int, str],
    ref_errors: List[Dict[str, Optional[str]]],
    ref_seen: Set[str],
) -> Dict[str, Optional[str]]:
    due_dates: Dict[str, Optional[str]] = {}
    for col, bucket in bucket_cols.items():
        col_candidates = [col]
        if col == 4:  # Savings often stores the date in the adjacent column
            col_candidates.append(5)
        iso_value: Optional[str] = None
        for candidate_col in col_candidates:
            for row in range(start_row + 2, start_row + 6):
                cell_value, formula = _inspect_cell(
                    ws_values,
                    ws_formulas,
                    row,
                    candidate_col,
                    ref_errors,
                    ref_seen,
                )
                iso_value = _to_iso_date(cell_value)
                if iso_value:
                    break
            if iso_value:
                break
        due_dates[bucket] = iso_value
    return due_dates


def _is_next_block_start(ws, row: int, current_start: int) -> bool:
    if row <= current_start:
        return False
    cell_value = ws.cell(row=row, column=1).value
    header_flag = ws.cell(row=row, column=6).value
    return isinstance(cell_value, datetime) and isinstance(header_flag, str)


def _extract_accounts(
    ws_values,
    ws_formulas,
    start_row: int,
    bucket_cols: Dict[int, str],
    ref_errors: List[Dict[str, Optional[str]]],
    ref_seen: Set[str],
) -> Tuple[List[AccountAllocation], int]:
    accounts: List[AccountAllocation] = []
    row = start_row + 5  # first account row
    last_row = start_row + 4
    while row <= ws_values.max_row and not _is_next_block_start(ws_values, row, start_row):
        name_cell_value, _ = _inspect_cell(
            ws_values,
            ws_formulas,
            row,
            5,
            ref_errors,
            ref_seen,
        )
        name = name_cell_value
        if not isinstance(name, str) or not name.strip():
            row += 1
            continue

        remaining_raw, remaining_formula = _inspect_cell(
            ws_values, ws_formulas, row, 1, ref_errors, ref_seen
        )
        fixed_raw, fixed_formula = _inspect_cell(
            ws_values, ws_formulas, row, 2, ref_errors, ref_seen
        )
        savings_raw, savings_formula = _inspect_cell(
            ws_values, ws_formulas, row, 4, ref_errors, ref_seen
        )

        bucket_allocations: Dict[str, Number] = {}
        bucket_formulas: Dict[str, Optional[str]] = {}
        for col, bucket in bucket_cols.items():
            if col == 5:
                continue
            bucket_raw, bucket_formula = _inspect_cell(
                ws_values,
                ws_formulas,
                row,
                col,
                ref_errors,
                ref_seen,
            )
            bucket_allocations[bucket] = _to_number(bucket_raw)
            bucket_formulas[bucket] = bucket_formula

        accounts.append(
            AccountAllocation(
                name=name.strip(),
                remaining_cash=_to_number(remaining_raw),
                fixed_balance=_to_number(fixed_raw),
                savings_transfer=_to_number(savings_raw),
                bucket_allocations=bucket_allocations,
                formulas={
                    "remaining_cash": remaining_formula,
                    "fixed_balance": fixed_formula,
                    "savings_transfer": savings_formula,
                },
                bucket_formulas=bucket_formulas,
            )
        )
        row += 1
        last_row = row - 1

    return accounts, last_row


def _parse_month_block(
    ws_values,
    ws_formulas,
    start_row: int,
) -> PlannedMonthExport:
    month_start_cell = ws_values.cell(row=start_row, column=1).value
    if not isinstance(month_start_cell, datetime):
        raise ValueError(f"Row {start_row} does not contain a valid month start date.")

    ref_errors: List[Dict[str, Optional[str]]] = []
    ref_seen: Set[str] = set()

    legend_row = start_row + 4
    bucket_cols = _extract_bucket_columns(ws_values, legend_row)
    status_row = start_row + 1
    statuses = _extract_statuses(ws_values, status_row, bucket_cols)

    due_dates = _extract_due_dates(
        ws_values,
        ws_formulas,
        start_row,
        bucket_cols,
        ref_errors,
        ref_seen,
    )
    accounts, end_row = _extract_accounts(
        ws_values,
        ws_formulas,
        start_row,
        bucket_cols,
        ref_errors,
        ref_seen,
    )

    fixed_factor_raw, fixed_factor_formula = _inspect_cell(
        ws_values,
        ws_formulas,
        start_row + 2,
        2,
        ref_errors,
        ref_seen,
    )
    inflow_raw, inflow_formula = _inspect_cell(
        ws_values,
        ws_formulas,
        start_row + 3,
        1,
        ref_errors,
        ref_seen,
    )

    month = PlannedMonthExport(
        month_start=month_start_cell.date().isoformat(),
        fixed_factor=_to_number(fixed_factor_raw),
        inflow_total=_to_number(inflow_raw),
        inflow_formula=inflow_formula,
        fixed_factor_formula=fixed_factor_formula,
        status_by_bucket=statuses,
        due_dates=due_dates,
        bucket_order=list(bucket_cols.values()),
        accounts=accounts,
        source_rows={"start": start_row, "end": end_row},
        ref_errors=ref_errors,
    )

    return month


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export Planned Expenses sheet to JSON.")
    parser.add_argument(
        "--workbook",
        default="docs/Copy of Expense Shares.xlsx",
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--sheet",
        default="Planned Expenses",
        help="Worksheet name containing the monthly planner.",
    )
    parser.add_argument(
        "--output",
        default="data/seeds/planned-expenses.json",
        help="Path to write JSON output.",
    )
    parser.add_argument(
        "--limit-months",
        type=int,
        default=None,
        help="Optional limit of months to export (starting from the top).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    workbook_path = (ROOT_DIR / args.workbook).resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb_values = load_workbook(workbook_path, data_only=True)
    wb_formulas = load_workbook(workbook_path, data_only=False)

    if args.sheet not in wb_values.sheetnames:
        raise SystemExit(f"Worksheet '{args.sheet}' not found. Available: {wb_values.sheetnames}")
    ws_values = wb_values[args.sheet]
    ws_formulas = wb_formulas[args.sheet]

    months: List[PlannedMonthExport] = []
    for idx, start_row in enumerate(_iter_month_starts(ws_values), start=1):
        months.append(_parse_month_block(ws_values, ws_formulas, start_row))
        if args.limit_months is not None and idx >= args.limit_months:
            break

    output_path = (ROOT_DIR / args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    serialisable = [asdict(month) for month in months]
    with output_path.open("w", encoding="utf-8") as fh:
        json.dump(serialisable, fh, indent=2, ensure_ascii=False)

    print(f"Exported {len(serialisable)} month(s) to {output_path}")
    ref_count = sum(len(month.ref_errors) for month in months)
    if ref_count:
        print(f"Detected {ref_count} cell(s) containing #REF! or Excel errors.")


if __name__ == "__main__":  # pragma: no cover
    main()

